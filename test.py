from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl import Workbook
import configparser
import csv

from get_data import well_compound_list, get_survey_csv_data, get_all_trans_data
from get_data import get_xml_trans_data_skipping_wells


def tester(config, file):
    all_data = {}


    with open(file, "r") as log:
        lines = log.readlines()

        for line in lines:
            info = line.split()
            try:
                info[0]
            except IndexError:
                pass
            else:
                if info[0] == "Start":

                    Skip_data_point = False
                    for counter, data in enumerate(info):
                        if data == "Time:":
                            temp_date = info[counter + 1]
                            temp_time = info[counter + 2]
                        elif data == "Test#:":
                            temp_test = info[counter + 1]
                        elif data == "Device:":
                            temp_device = info[counter + 1]
                        elif data == "Method:":
                            temp_device_method = info[counter + 1]
                        elif data == "Id:":
                            temp_id = info[counter + 1]
                        elif data == "Command:":
                            if info[counter + 1] == "VISIT":
                                Skip_data_point = True

                    if not Skip_data_point and temp_id != "None" and temp_test != "2":
                        try:
                            all_data[temp_id]
                        except KeyError:
                            all_data[temp_id] = {}

                        try:
                            all_data[temp_id][temp_device]
                        except KeyError:
                            all_data[temp_id][temp_device] = {}

                        try:
                            all_data[temp_id][temp_device][temp_device_method]
                        except KeyError:
                            temp_counter = 1
                            all_data[temp_id][temp_device][temp_device_method] = {temp_counter: {"start_time": temp_time,
                                                                                            "start_date": temp_date,
                                                                                             "end_time": "",
                                                                                             "end_date": ""}}

                        else:

                            for numbers in all_data[temp_id][temp_device][temp_device_method]:
                                temp_counter = numbers + 1
                            all_data[temp_id][temp_device][temp_device_method] = {temp_counter:
                                                                                            {"start_time": temp_time,
                                                                                            "start_date": temp_date,
                                                                                             "end_time": "",
                                                                                             "end_date": ""}}
                        temp_id_place_holder = temp_id
                        temp_device_place_holder = temp_device
                        temp_device_method_place_holder = temp_device_method
                        temp_counter_place_holder = temp_counter

                if info[0] == "End":


                    for counter, data in enumerate(info):
                        if data == "Time:":
                            temp_date = info[counter + 1]
                            temp_time = info[counter + 2]
                        elif data == "Id:":
                            temp_id = info[counter + 1]

                    if temp_id != "None":

                        all_data[temp_id_place_holder][temp_device_place_holder][temp_device_method_place_holder][temp_counter_place_holder]["end_time"] = temp_time
                        all_data[temp_id_place_holder][temp_device_place_holder][temp_device_method_place_holder][temp_counter_place_holder]["end_date"] = temp_date

    wb = Workbook()
    ws = wb.active

    row = 1
    col = 1

    for plates in (all_data):
        ws.cell(row=row, column=col, value=plates)
        col += 1
        for devices in all_data[plates]:
            for method in all_data[plates][devices]:
                for counter in all_data[plates][devices][method]:
                    ws.cell(row=row, column=col + 0, value=devices)
                    ws.cell(row=row, column=col + 1, value=method)
                    ws.cell(row=row, column=col + 2, value=all_data[plates][devices][method][counter]["start_time"])
                    ws.cell(row=row, column=col + 3, value=all_data[plates][devices][method][counter]["end_time"])
                col += 4
        col = 1
        row += 1
    wb.save("C:/Users/phch/Desktop/more_data_files/log_book.xlsx")



if __name__ == "__main__":
    config = configparser.ConfigParser()
    config.read("config.ini")



    trans_data_folder = "C:/Users/phch/Desktop/more_data_files/2022-11-22"
    plate_layout_folder = "C:/Users/phch/Desktop/more_data_files/plate_layout"
    data_location = "C:/Users/phch/Desktop/more_data_files/2022-11-22"
    file_name = "test_trans_report"
    save_location = "C:/Users/phch/Desktop/more_data_files/"
    all_trans_file = "C:/Users/phch/Desktop/more_data_files/all_trans.xlsx"
    survey_folder = "C:/Users/phch/Desktop/more_data_files/surveys"
    path = "C:/Users/phch/Desktop/echo_data"
    save_file_name = "test_new_worklist"
    file = "C:/Users/phch/Desktop/more_data_files/PlateButler.log"
    dead_vol_ul = {"LDV": 2.5, "PP": 15.0}

    set_amount = 1

    tester(config, file)

    # new_worklist(survey_folder, plate_layout_folder, all_trans_file, set_amount, dead_vol_ul, save_location, save_file_name)
    # excel_controller(trans_data_folder, plate_layout_folder, all_trans_file, data_location, file_name, save_location)
    # get_comments(all_trans_file)

    # print(file_names(path))

