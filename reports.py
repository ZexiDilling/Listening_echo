import PySimpleGUI
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl import Workbook, load_workbook

from get_data import get_xml_trans_data_skipping_wells, get_xml_trans_data_printing_wells, well_compound_list,\
    get_survey_csv_data, get_all_trans_data
import natsort


def _write_to_excel_plate_transferees_list_of_plates(wb, data, title):
    """
    Writes witch source plates goes to witch destination plate
    :param wb: excel Workbook
    :type wb: openpyxl.workbook.workbook.Workbook
    :param data: Data with all plate transferees
    :type data: dict
    :return:
    """

    # Create a new worksheet and give it a title
    ws = wb.create_sheet(title)

    # Write the header row
    header_destination = "destination plate"
    header_source = "source plates"
    ws.cell(row=1, column=1, value=header_destination).font = Font(bold=True)
    ws.merge_cells("B1:F1")
    ws["B1"].value = header_source
    ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B1"].font = Font(bold=True)

    # Initialize variables to keep track of rows and columns
    col = 1
    row = 2

    # Write the data to the worksheet
    for destination in data:
        ws.cell(row=row, column=col, value=destination)
        for index, source in enumerate(data[destination]):
            ws.cell(row=row, column=col + 1 + index, value=source)

        # Increment the row for the next destination
        row += 1

def _write_to_excel_plate_transferees(wb, data, title):    # TODO description
    """
    Writes the destination plate and the count of source plates transfered to it
    :param wb: excel Workbook
    :type wb: openpyxl.workbook.workbook.Workbook
    :param data: Data with all plate transferees count
    :type data: list
    :param title: Title of the sheet
    :type title: str
    """
    ws = wb.create_sheet(title)

    # Write headers
    ws.cell(row=1, column=1, value="destination plate").font = Font(bold=True)
    ws.cell(row=1, column=2, value="counter").font = Font(bold=True)

    # Write the data, barcodes for destination plates and the number of transferees
    for row, plate_counts in enumerate(data, 2):
        plate, counts = plate_counts.split(",")
        ws.cell(row=row, column=1, value=plate)
        ws.cell(row=row, column=2, value=counts)


def _write_to_excel_error_report(wb, data, title): # TODO description
    """
    Writes the summary of the plate transferees data to excel.
    :param wb: excel Workbook
    :type wb: openpyxl.workbook.workbook.Workbook
    :param data: Data with all plate transferees information.
    :type data: list
    :param title: Sheet title
    :type title: str
    """

    # Create the sheet
    ws = wb.create_sheet(title)

    # Start writing the data
    row = 1
    col = 1
    i = 0
    while i < len(data):

        # Write the date, source barcode, destination barcode, and amount of wells skipped as headers
        ws.cell(row=row, column=col, value=data[i + 1]).font = Font(bold=True)  # date
        ws.cell(row=row, column=col + 1, value=data[i + 2]).font = Font(bold=True)  # source barcode
        ws.cell(row=row, column=col + 2, value=data[i + 3]).font = Font(bold=True)  # destination barcode
        ws.cell(row=row, column=col + 3, value=data[i]).font = Font(bold=True)  # writes amount of wells skipped

        n = data[i]  # Get the number of wells skipped
        temp = i + 4
        i = temp
        q = 0

        # Write the data for the wells skipped
        for k in range(temp, (int(n) * 2) + temp, 2):

            if q == 5:  # Write to next row after writing to 5 columns
                row = row + 2
                q = 0

            ws.cell(row=row + 1, column=col + q, value=data[k])  # Write the source well
            ws.cell(row=row + 2, column=col + q, value=data[k + 1])  # Write the destination well
            q += 1

        row = row + 3  # Increment the row number for the next section
        col = col  # Keep the column number constant
        i = int(n) * 2 + i  # Increment the index for the next section


def _write_work_list(wb, data, title):
    """
    This function writes data to a worksheet in an existing workbook (wb) with the given title (title)
    :param wb: existing workbook object
    :type wb: Workbook
    :param data: data to be written to the worksheet
    :type data: dict
    :param title: title of the worksheet to be created
    :type title: str
    """
    # create a new worksheet with the given title
    ws = wb.create_sheet(title)

    # initialize row and column indices
    row = 1
    col = 1

    # write headlines
    ws.cell(row=row, column=col, value="source_plate")
    ws.cell(row=row, column=col + 1, value="source_well")
    ws.cell(row=row, column=col + 2, value="trans_vol")
    ws.cell(row=row, column=col + 3, value="destination_well")
    ws.cell(row=row, column=col + 4, value="destination_plate")

    # increment the row index to move to the next row
    row += 1

    # loop through each destination plate in the data
    for destination in data:
        # loop through each source plate in the data for the current destination
        for source in data[destination]:
            # loop through each transfer in the data for the current source and destination
            for index, trans in enumerate(data[destination][source]):
                # write the source plate to the current row and column
                ws.cell(row=row, column=col, value=source)

                # loop through the source well, transfer volume, and destination well information
                for off_set, info in enumerate(data[destination][source][index]):
                    # write the source well, transfer volume, and destination well information to the current row and columns
                    ws.cell(row=row, column=col + off_set + 1, value=str(info))

                # write the destination plate to the current row and column
                ws.cell(row=row, column=col + 4, value=destination)

                # increment the row index to move to the next row
                row += 1


def _skip_report(wb, skipped_wells, skip_well_counter, working_list, config, overview_data, title):
    """
    Extracts data from an excel workbook and appends to working list. Skips wells in skipped_wells.
    :param wb: excel Workbook
    :type wb: openpyxl.workbook.workbook.Workbook
    :param skipped_wells: List of skipped well names
    :type skipped_wells: List[str]
    :param skip_well_counter: Count of skipped wells
    :type skip_well_counter: int
    :param working_list: List of wells to process
    :type working_list: List[str]
    :param config: The config handler, with all the default information in the config file.
    :type config: configparser.ConfigParser
    :param overview_data: Data for the overview sheet
    :type overview_data: dict
    :param title: Title of the workbook
    :type title: str
    :return: None
    """

    ws = wb.active
    ws.title = title
    row = 1
    col = 1
    temp_reason = ""

    source_plate_amount = 0
    destination_plate_amount = 0
    source_plate_list = []
    failed_plates = []
    for destination_plates in working_list:
        destination_plate_amount += 1
        failed_plates.append(destination_plates)
        for source_plate in working_list[destination_plates]:

            if source_plate not in source_plate_list:
                source_plate_list.append(source_plate)
                source_plate_amount += 1

    # overview
    ws.cell(row=row, column=col, value="Skipped Wells counter").font = Font(bold=True)
    ws.cell(row=row, column=col + 1, value=skip_well_counter)
    row += 1
    ws.cell(row=row, column=col, value="Source Plates counter").font = Font(bold=True)
    ws.cell(row=row, column=col + 1, value=source_plate_amount)
    row += 1
    ws.cell(row=row, column=col, value="Destination plate counter").font = Font(bold=True)
    ws.cell(row=row, column=col + 1, value=destination_plate_amount)
    row += 1
    ws.cell(row=row, column=col, value="Plate transfers").font = Font(bold=True)
    ws.cell(row=row, column=col + 1, value=destination_plate_amount)
    row += 2
    ws.cell(row=row, column=col, value="Destination Plates").font = Font(bold=True)
    ws.cell(row=row, column=col + 1, value="Source Plates").font = Font(bold=True)
    row += 1


    temp_row = row
    for failed_plate in failed_plates:
        ws.cell(row=temp_row, column=col, value=failed_plate)
        temp_row += 1

    temp_row = row
    for source in source_plate_list:
        ws.cell(row=temp_row, column=col + 1, value=source)
        temp_row += 1

    temp_row = row
    temp_col = col
    last_row = 0
    failed_source_wells = 0
    for source_plate in skipped_wells:
        ws.cell(row=temp_row - 1, column=temp_col + 3, value="Source Plates").font = Font(bold=True)
        ws.cell(row=temp_row - 1, column=temp_col + 4, value="Source Wells").font = Font(bold=True)
        ws.cell(row=temp_row - 1, column=temp_col + 5, value="Well Instance").font = Font(bold=True)
        ws.cell(row=temp_row - 1, column=temp_col + 6, value="Well Volume").font = Font(bold=True)
        ws.cell(row=temp_row - 1, column=temp_col + 7, value="Reason").font = Font(bold=True)

        ws.cell(row=temp_row, column=temp_col + 3, value=source_plate)
        for index, source_well in enumerate(skipped_wells[source_plate]):
            ws.cell(row=temp_row + index, column=temp_col + 4, value=source_well)
            ws.cell(row=temp_row + index, column=temp_col + 5, value=skipped_wells[source_plate][source_well]["counter"])
            ws.cell(row=temp_row + index, column=temp_col + 6, value=skipped_wells[source_plate][source_well]["vol"])

            # sort out reason. remove duplicate in the list, and writes them all in one cell
            temp_reason_list = list(set(skipped_wells[source_plate][source_well]["reason"]))
            for reasons in temp_reason_list:
                temp_reason += f"{reasons}, "
            temp_reason = temp_reason.removesuffix(", ")
            try:
                colour = config["Echo_error_colours"][temp_reason]
            except KeyError:
                colour = "c3ded6"

            ws.cell(row=temp_row + index, column=temp_col + 7, value=temp_reason).fill = PatternFill(start_color=colour,
                                                                                     end_color=colour,
                                                                                     fill_type='solid')  # Error code

            failed_source_wells += 1
            temp_reason = ""
            if temp_row + index > last_row:
                last_row = temp_row + index

        temp_col += 5

    # Writes translation for errors:
    row = last_row + 2
    col = 4

    # Headlines:
    ws.cell(row=row, column=col, value="Code").font = Font(bold=True)
    ws.cell(row=row, column=col + 1, value="Echo Reason").font = Font(bold=True)
    ws.cell(row=row, column=col + 2, value="Translated Reason").font = Font(bold=True)

    row += 1

    for data_index, data in enumerate(config["Echo_error_real"]):
        try:
            colour = config["Echo_error_colours"][data]
        except KeyError:
            colour = "c3ded6"
        ws.cell(row=row + data_index, column=col, value=data).fill = PatternFill(start_color=colour,
                                                                                 end_color=colour,
                                                                                 fill_type='solid')  # Error code
        ws.cell(row=row + data_index, column=col + 1, value=config["Echo_error_real"][data])    # Echo msg
        ws.cell(row=row + data_index, column=col + 2, value=config["Echo_error"][data])     # Charlies guess as to why

    overview_data["amount_failed_plates"] = destination_plate_amount
    overview_data["failed_trans"] = skip_well_counter
    overview_data["amount_source_plates"] = source_plate_amount
    overview_data["failed_wells"] = failed_source_wells

    return failed_plates


def _write_completed_plates(wb, zero_error_trans_plate, failed_plates, overview_data, title):
    """
    Writes a list of completed plates, failed plates, and completed sets.
    :param wb: the workbook
    :param zero_error_trans_plate: A list of plates with zero errors
    :type zero_error_trans_plate: list
    :param failed_plates: A list of plates with errors
    :type failed_plates: list
    :param overview_data: a dict of overview date for the run:
    :type overview_data: dict
    :param title: The title of the excel sheet
    :type title: str
    :return: data writen into the overview_data dict
    """
    ws = wb.create_sheet(title)
    row = 1
    col = 1
    temp_plate_list = []
    complete_sets = []

    # Headlines:
    ws.cell(row=row, column=col, value="Complete plates").font = Font(bold=True)
    ws.cell(row=row, column=col, value="failed plates").font = Font(bold=True)
    ws.cell(row=row, column=col, value="complete sets").font = Font(bold=True)

    row += 1
    if zero_error_trans_plate:
        zero_error_trans_plate = natsort.natsorted(list(set(zero_error_trans_plate)))
        for plate_index, plates in enumerate(zero_error_trans_plate):
            ws.cell(row=row + plate_index, column=col, value=plates)
            plate = plates.split("-")[0]
            temp_plate_list.append(plate)
    else:
        ws.cell(row=row, column=col, value="All plates failed somewhere").font = Font(bold=True)

    if failed_plates:
        temp_plate_list = natsort.natsorted(list(set(temp_plate_list)))
        for plate_index, plates in enumerate(failed_plates):
            ws.cell(row=row + plate_index, column=col + 1, value=plates)
            plate = plates.split("-")[0]
            if plate not in temp_plate_list:
                complete_sets.append(f"{plate}-set")
    else:
        ws.cell(row=row, column=col + 1, value="All plates completed with zero errors").font = Font(bold=True)

    if complete_sets:
        complete_sets = natsort.natsorted(list(set(complete_sets)))
        for set_index, sets in enumerate(complete_sets):
            ws.cell(row=row + set_index, column=col + 2, value=sets)
    else:
        ws.cell(row=row, column=col + 2, value="No fully completed sets").font = Font(bold=True)

    overview_data["amount_complete_plates"] = len(zero_error_trans_plate)


def skipped_well_controller(data_location, full_path, config): # TODO description
    """
    Makes an excel sheet with data from the Echo.
    :param data_location: Where all the echo data is located
    :type data_location: str
    :param full_path: name of the final report, and location of where to save it.
    :type full_path: str
    :return: an overview of the run
    :rtype: dict
    """
    # Getting data for wells being skipped
    all_data, skipped_wells, skip_well_counter, working_list, trans_plate_counter, all_trans_counter, \
        zero_error_trans_plate, _ = get_xml_trans_data_skipping_wells(data_location)

    # Gets data from all transfers
    trans_data = get_xml_trans_data_printing_wells(data_location)

    # Create a dict with data to get a quick overview
    overview_data = {"plate_amount": 0,
                     "amount_complete_plates": 0,
                     "amount_failed_plates": 0,
                     "failed_wells": 0,
                     "failed_trans": 0,
                     "amount_source_plates": 0,
                     "time_for_all_trans": 0,
                     "path": ""}

    # Create the workbook
    wb = Workbook()

    # Add different reports to their own worksheet in an excel ark,
    # to create a full report over transfers and skipped wells
    failed_plates = _skip_report(wb, skipped_wells, skip_well_counter, working_list, config, overview_data, title="Overview_Report")
    _write_to_excel_plate_transferees_list_of_plates(wb, all_trans_counter, title="Plate_trans_list")
    _write_to_excel_plate_transferees(wb, trans_plate_counter, title="Plate_trans_counter")
    _write_to_excel_error_report(wb, all_data, title="Error_Report")
    _write_work_list(wb, working_list, title="Old_Worklist")
    _write_completed_plates(wb, zero_error_trans_plate, failed_plates, overview_data, title="Completed_plates")
    _write_trans_report(wb, trans_data, title="Trans_Report", compound_data=None, report=True)

    wb.save(full_path)

    overview_data["plate_amount"] = overview_data["amount_complete_plates"] + overview_data["amount_failed_plates"]
    overview_data["path"] = full_path

    return overview_data


def _rename_source_plates(trans_data, prefix_dict):
    """
    This function updates the source plate name in the `trans_data` dictionary by adding a prefix based on the date.
    :param trans_data: a dictionary containing the transaction data with keys as transaction IDs and values as a
        dictionary with keys "source_plate" and "date".
    :type trans_data: dict
    :param prefix_dict: a dictionary containing the prefix to be added to the source plate name, with keys as prefixes
        and values as a dictionary with keys "start" and "end", representing the start and end dates for the prefix.
    :type prefix_dict: dict
    :return: None
    """
    # Iterate over the transaction data
    for trans in trans_data:
        # Get the source plate name and date for the current transaction
        temp_source_plate = trans_data[trans]["source_plate"]
        temp_date = trans_data[trans]["date"]
        # Iterate over the prefix dictionary
        for prefix in prefix_dict:
            # Check if the date for the current transaction falls within the start and end dates for the prefix
            if prefix_dict[prefix]["start"] <= temp_date <= prefix_dict[prefix]["end"]:
                # Update the source plate name by adding the prefix
                trans_data[trans]["source_plate"] = f"{prefix}_{temp_source_plate}"


def _write_trans_report(wb, trans_data, compound_data, title, report): # TODO description

    # sets row and coloumn counter for headlines
    row = 1
    col = 1

    #Check what the function is being used for. If it is for the analysis, then there is no reason to add compound data
    #It needs to create a new worksheet if it is for the report.
    #Spacer is to make space for the 'compound' data if it is included
    if report:
        ws = wb.create_sheet(title)
        spacer = 0
    else:
        ws = wb.active
        ws.title = title
        spacer = 1
        ws.cell(row=row, column=col + 2, value="Compound").font = Font(bold=True)

    # Headers
    ws.cell(row=row, column=col + 0, value="Destination Plates").font = Font(bold=True)
    ws.cell(row=row, column=col + 1, value="Destination Well").font = Font(bold=True)
    ws.cell(row=row, column=col + spacer + 2, value="Volume").font = Font(bold=True)
    ws.cell(row=row, column=col + spacer + 3, value="Source Plates").font = Font(bold=True)
    ws.cell(row=row, column=col + spacer + 4, value="Source Well").font = Font(bold=True)

    row += 1

    for row_index, trans in enumerate(trans_data):
        destination_plate = trans_data[trans]["destination_plate"]
        source_plate = trans_data[trans]["source_plate"]
        volume = trans_data[trans]["transferees"]["volume"]
        destination_well = trans_data[trans]["transferees"]["destination_well"]
        source_well = trans_data[trans]["transferees"]["source_well"]
        if compound_data:
            compound = compound_data[source_plate][source_well]

        ws.cell(row=row + row_index, column=col + 0, value=destination_plate)   #.font = Font(bold=True)
        ws.cell(row=row + row_index, column=col + 1, value=destination_well)    #.font = Font(bold=True)
        ws.cell(row=row + row_index, column=col + spacer + 2, value=volume)  #.font = Font(bold=True)
        ws.cell(row=row + row_index, column=col + spacer + 3, value=source_plate)    #.font = Font(bold=True)
        ws.cell(row=row + row_index, column=col + spacer + 4, value=source_well) #.font = Font(bold=True)
        if not report and compound_data:
            ws.cell(row=row + row_index, column=col + 2, value=compound)  # .font = Font(bold=True)


def trans_report_controller(trans_data_folder, plate_layout_folder, file_name, save_location):
    """
    Generates a trans report in excel format and saves it to a specified location.

    :param trans_data_folder: The path to the folder containing trans data.
    :type trans_data_folder: str
    :param plate_layout_folder: The path to the folder containing plate layout information.
    :type plate_layout_folder: str
    :param file_name: The desired name of the generated excel file.
    :type file_name: str
    :param save_location: The path to the location where the generated file should be saved.
    :type save_location: str
    :return: A message indicating the completion of the operation.
    :rtype: str
    """

    # Get the trans data and well compound list
    trans_data = get_xml_trans_data_printing_wells(trans_data_folder)
    compound_data = well_compound_list(plate_layout_folder)

    # Set the save file location
    save_file = f"{save_location}/{file_name}.xlsx"

    # Check if prefixing is enabled
    prefix_on = True
    prefix_dict = {"OLD": {"start": "2022-11-22", "end": "2022-12-01"},
              "NEW": {"start": "2022-10-01", "end": "2022-11-21"}}
    if prefix_on:
        _rename_source_plates(trans_data, prefix_dict)

    # Create a workbook
    wb = Workbook()

    # Write the trans report to the workbook
    _write_trans_report(wb, trans_data, compound_data, title="Trans_report", report=False)

    # Save the workbook
    wb.save(save_file)

    # Return the completion message
    return "Done"


def well_report(trans_file, save_file):
    """
    Generates a report of source wells and their count, volume, and compound.
    :param trans_file:path to the source excel file
    :type trans_file str
    :param save_file: path to the destination excel file
    :type save_file: str
    :return: None
    """

    # Load workbook

    wb = load_workbook(trans_file)
    ws = wb.active

    all_wells = {}

    for row_index, row in enumerate(ws.iter_rows(values_only=True)):
        if row_index == 0:
            continue
        comment, destination_plate, destination_well, compound, volume, source_well, source_plate, source_plate_type = row
        all_wells.setdefault(source_plate, {}).setdefault(source_well, {"counter": 0, "volume": 0, "compound": ""})
        all_wells[source_plate][source_well]["counter"] += 1
        all_wells[source_plate][source_well]["volume"] += float(volume)
        all_wells[source_plate][source_well]["compound"] = compound

    # Create new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Well Report"

    # Headers
    headers = ["Source Plates", "Source Well", "count", "volume", "compound"]
    for temp_col, header in enumerate(headers, 1):
        ws.cell(row=1, column=temp_col, value=header).font = Font(bold=True)

    row = 1
    col = 1

    for plates, wells in all_wells.items():
        for well, data in wells.items():
            ws.cell(row=row, column=col + 0, value=plates)
            ws.cell(row=row, column=col + 1, value=well)
            ws.cell(row=row, column=col + 2, value=data["counter"])
            ws.cell(row=row, column=col + 3, value=data["volume"])
            ws.cell(row=row, column=col + 4, value=data["compound"])
            row += 1

    wb.save(save_file)
    print("done")


def _compound_to_survey(plate_layout, survey_data):
    """
    This function maps compounds to a survey layout.

    :param plate_layout: A dictionary containing information about the compounds present in each well of a plate.
    :type plate_layout: dict
    :param survey_data: A dictionary containing the volume of a compound in each well of a plate.
    :type survey_data: dict
    :return: A dictionary that maps compounds to the survey layout.
    :rtype: dict
    """

    survey_layout = {}

    # loop through each plate in the survey data
    for plate in survey_data:
        # create a new key in the survey layout dictionary for the current plate, if not already present
        if plate not in survey_layout:
            survey_layout[plate] = {}

        # loop through each plate name in the survey data for the current plate
        for plate_name in survey_data[plate]:
            # loop through each well in the survey data for the current plate name
            for well in survey_data[plate][plate_name]:
                # if the volume for the well is not 0
                if survey_data[plate][plate_name][well] != 0:
                    # get the compound name from the plate layout for the current well
                    try:
                        compound = plate_layout[plate_name][well]
                    except KeyError:
                        # if there is no matching compound, set it as "No compound match found"
                        compound = "No compound match found"

                    # get the volume for the current well
                    volume = survey_data[plate][plate_name][well]

                    # create a new key in the survey layout dictionary for the current compound, if not already present
                    if compound not in survey_layout[plate]:
                        survey_layout[plate][compound] = {}

                    # create a new key in the survey layout dictionary for the current plate name,
                    # if not already present
                    if plate_name not in survey_layout[plate][compound]:
                        survey_layout[plate][compound][plate_name] = {}

                    # set the value for the well to the volume
                    survey_layout[plate][compound][plate_name][well] = float(volume)

    return survey_layout


def _write_new_worklist(set_compound_data, survey_layout, dead_vol_ul, set_amount, save_file, starting_set,
                        specific_transfers):
    """
    writes two worklist for plate_printing.
    1 for LDV trans
    1 for PP trans
    and sets them in one excel sheet
    :param set_compound_data: Data over compounds
    :type set_compound_data: dict
    :param survey_layout: combination of survey data for a plate, and the layout of the plate. To know how much is left
        of each compound
    :type survey_layout: dict
    :param dead_vol_ul: dead volume for plates. There are defaults in the config.
    :type dead_vol_ul: dict
    :param set_amount: Amount of sets to produce. This do not take into account the starting set. so set it to 50 and
        starting set at 40, and it will only produce 10 sets.
    :type set_amount: int
    :param save_file: Where to save the data and what name to call it
    :type save_file: str
    :param starting_set: The number of the first set. To make it possible to start from a different number than 1
    :type starting_set: int
    :param specific_transfers: Takes a dict over specific transfers that needs to be made. In case some plates are
        ruined for a run.
    :type specific_transfers: dict
    :return: an excel file with data

    """
    wb = Workbook()
    # Create a sheet for each transferee, as we can't run LDV and PP trans at the same time
    ws0 = wb.active
    ws1 = wb.create_sheet("LDV_trans")
    ws2 = wb.create_sheet("PP_trans")

    # Failed plates, are plates not enough liquid for a transferee. name is depending on plate name and compound
    failed_plate = {}

    # Compound dict is a dict of compounds for specific plates, where dead volume is set lower, due to missing liquids.
    # An Echo will mostly work under "dead volume".
    compound_list = {"P23_LDV": {"Buparlisib": 2.0, "BGB-11417": 2.2}, "P23_LDV2_I2": {"Copanlisib": 2.0}}
    # compound_list = {}

    # Breaking is used, to make sure that all wells are used. instead of all transfers being using the last well in
    # the list
    breaking = False

    # Setup rows for the different worksheets
    row_ldv = 1
    row_pp = 1
    row = 1
    col = 1

    # setup error msg.
    error_missing_survey_data = "Missing survey Data"
    error_missing_liquid = "Not enough liquid"

    # headlines:
    headlines = ["source_plates", "source_well", "volume", "destination_well", "destination_plates", "compound"]

    for temp_col, headline in enumerate(headlines):
        ws1.cell(row=row, column=col + temp_col, value=headline).font = Font(bold=True)    # headers_LDV:
        ws2.cell(row=row, column=col + temp_col, value=headline).font = Font(bold=True)    # headers_PP:


    if not starting_set:
        starting_set = 1

    if specific_transfers:
        plate_range = specific_transfers

    else:
        plate_range = range(set_amount + 1 - starting_set)

    for sets in plate_range:
        for rows in set_compound_data:
            plate_letter = set_compound_data[rows]['destination_plate']
            if specific_transfers:
                if plate_letter[-1] != sets.split("-")[-1]:
                    continue
                destination_plate = sets
            else:
                destination_plate = f"{sets}-{plate_letter}"
            destination_well = set_compound_data[rows]["destination_well"]
            compound = set_compound_data[rows]["compound"]
            volume_nl = float(set_compound_data[rows]["volume_nl"])
            volume_ul = volume_nl/1000
            # sample_comment = set_compound_data[rows]["sample_comment"]
            source_plate_origin = set_compound_data[rows]["source_plate"]
            if specific_transfers:
                if "pp" in source_plate_origin.casefold():
                    temp_plate_type = "PP"
                elif "ldv" in source_plate_origin.casefold():
                    temp_plate_type = "LDV"
                else:
                    continue
                if not plate_range[sets][temp_plate_type]:
                    continue
            plate_type = set_compound_data[rows]["plate_type"]

            # Change between LDV and PP trans, depending on the source plate type
            if "pp" in source_plate_origin.casefold():
                ws = ws2
                row_pp += 1
                row = row_pp
            else:
                ws = ws1
                row_ldv += 1
                row = row_ldv

            # Check if there is survey data for the plate, and for the compound in the plate
            try:
                survey_layout[source_plate_origin][compound]
            except KeyError:
                source_plate = source_plate_origin
                source_well = error_missing_survey_data
            else:

                # Break is here to break out of the loop, to avoid looping over all the wells with the same compound in
                # them, for all the plate where the compound is.
                for temp_plates in survey_layout[source_plate_origin][compound]:
                    for wells in survey_layout[source_plate_origin][compound][temp_plates]:
                        try:
                            compound_list[source_plate_origin]
                        except KeyError:
                            dead_vol = dead_vol_ul[plate_type]
                        else:
                            if compound in compound_list[source_plate_origin]:
                                dead_vol = compound_list[source_plate_origin][compound]
                            else:
                                dead_vol = dead_vol_ul[plate_type]

                        if survey_layout[source_plate_origin][compound][temp_plates][wells] >= volume_ul + dead_vol:
                            source_plate = temp_plates
                            source_well = wells
                            survey_layout[source_plate_origin][compound][temp_plates][wells] -= volume_ul
                            breaking = True
                            break
                        else:
                            source_plate = f"{source_plate_origin}"
                            source_well = error_missing_liquid

                    if breaking:
                        breaking = False
                        break

            # This will colour cells where there is an error, and colour the headlines, to ensure notability.
            if source_well == error_missing_survey_data or source_well == error_missing_liquid:
                ws.cell(row=row, column=col + 0, value=f"{source_plate}_{compound}").fill = PatternFill(start_color='B284BE',
                                                                                        end_color='B284BE',
                                                                                        fill_type='solid')
                ws.cell(row=row, column=col + 1, value=source_well).fill = PatternFill(start_color='B284BE',
                                                                                       end_color='B284BE',
                                                                                       fill_type='solid')
                ws.cell(row=1, column=1).fill = PatternFill(start_color='B284BE', end_color='B284BE',
                                                                         fill_type='solid')
                ws.cell(row=1, column=2).fill = PatternFill(start_color='B284BE', end_color='B284BE',
                                                            fill_type='solid')
                failed_plate_compound = f"{source_plate}_{compound}"
                try:
                    failed_plate[failed_plate_compound]
                except KeyError:
                    failed_plate[failed_plate_compound] = {"plate": source_plate, "vol_ul": 0.0, "compound": compound, "trans_counter": 0}

                failed_plate[failed_plate_compound]["vol_ul"] += volume_ul
                failed_plate[failed_plate_compound]["trans_counter"] += 1

            # Writes the data for each transfer.
            else:
                ws.cell(row=row, column=col + 0, value=source_plate)
                ws.cell(row=row, column=col + 1, value=source_well)
            ws.cell(row=row, column=col + 2, value=volume_nl)
            ws.cell(row=row, column=col + 3, value=destination_well)
            ws.cell(row=row, column=col + 4, value=destination_plate)
            ws.cell(row=row, column=col + 5, value=compound)
            # ws.cell(row=row, column=col + 6, value=sample_comment)
            row += 1

    ws = ws0
    headlines_report = ["plates", "vol_ul", "compound", "trans_counter"]
    row = 1
    col = 1
    col_ldv = 1
    col_pp = col_ldv + len(headlines_report) + 1
    row_ldv = 2
    row_pp = row_ldv

    # Headlines
    for counter in range(2):
        for headlines in headlines_report:
            ws.cell(row=row, column=col, value=headlines)
            col += 1

    for index, plates in enumerate(failed_plate):
        if "LDV" in plates:
            row = row_ldv
            col = col_ldv
            row_ldv += 1

        if "PP" in plates:
            row = row_pp
            col = col_pp
            row_pp += 1

        ws.cell(row=row, column=col + 0, value=failed_plate[plates]["plate"])
        ws.cell(row=row, column=col + 1, value=failed_plate[plates]["vol_ul"])
        ws.cell(row=row, column=col + 2, value=failed_plate[plates]["compound"])
        ws.cell(row=row, column=col + 3, value=failed_plate[plates]["trans_counter"])

    wb.save(save_file)


def new_worklist(survey_folder, plate_layout_folder, file_trans, set_amount, dead_vol_ul, save_location,
                 save_file_name, specific_transfers=None, starting_set=None, window=None): # TODO description
    """
    This is the control function for making a new worklsit.
    :param survey_folder: The path to a folder where all the survey data is located
        The file names are the plate-barcode, the file-type is CSV, the layout is:
        row-1: Date
        row-2: Time
        row-3: empty
        row-4: well number
        row-5 -> last-row: letter + vol per well.
    :type survey_folder: str
    :param plate_layout_folder: The path to a folder where all the data for the plat layout is located.
        The file names are the plate-barcode, the file type is excel, col-1 is the well, col-2 is the drug/compound name
    :param file_trans: A file with all the transferes needed for at-least one full set. File-type is excel
    :type: str
    :param set_amount: amount of sets needed
    :type set_amount: int
    :param dead_vol_ul: dead volume for plates. There are defaults in the config.
    :type dead_vol_ul: dict
    :param save_location: Where the file is saved
    :type save_location: str
    :param save_file_name: What the file is called
    :type save_file_name: str
    :param specific_transfers: If there are specific plates that needs to be made, instaed of a full set
    :type specific_transfers: dict
    :param starting_set: What set to start at
    :type starting_set: int
    :param window: The GUI window
    :type window: PySimpleGUI.PySimpleGUI.Window
    :return: None
    """

    save_file = f"{save_location}/{save_file_name}.xlsx"

    survey_data = get_survey_csv_data(survey_folder)
    print("Got Survey Data")
    plate_layout = well_compound_list(plate_layout_folder)
    print("Got Plate layout")
    _, _, set_compound_data = get_all_trans_data(file_trans)
    print("Got compound data")
    survey_layout = _compound_to_survey(plate_layout, survey_data)
    print("got survey layout")
    _write_new_worklist(set_compound_data, survey_layout, dead_vol_ul, set_amount, save_file, starting_set, specific_transfers)
    if window:
        window["-WORKLIST_KILL-"].update(value=True)
    print("done")


# TODO Make report for compounds needed per set.... "Survey report" - burde ligge et sted.
# TODO make platlayout for 50 sets.

if __name__ == "__main__":
    pass
    # trans_data_folder = "C:/Users/phch/Desktop/more_data_files/2022-11-22"

    # data_location = "C:/Users/phch/Desktop/more_data_files/2022-11-22"
    # file_name = "test_trans_report"

    # all_trans_file = "C:/Users/Openscreen/Desktop/more_data_files/all_trans.xlsx"
    # path = "C:/Users/phch/Desktop/echo_data"


    # #
    # plate_layout_folder = "D:/plate_layout"
    # save_location = "C:/Users/phch/Desktop/more_data_files/"
    # file_trans = "D:/all_trans.xlsx"
    # set_amount = 50
    # starting_set = 17
    # dead_vol_ul = {"LDV": 3, "PP": 15}
    # save_file_name = "last_few_plates"
    #
    # survey_folder = "C:/Users/phch/Desktop/more_data_files/270123"
    #
    # plate_list = ["13-plate-C", "13-plate-D", "14-plate-B", "14-plate-C", "14-plate-D", "15-plate-C", "15-plate-D", "16-plate-C", "16-plate-D", "23-plate-C", "51-plate-C"]
    # specific_transfers = {}
    # for plate_names in plate_list:
    #     specific_transfers[plate_names] = {"LDV": True, "PP": False}
    #
    # print(specific_transfers)
    # #
    # #
    # new_worklist(survey_folder, plate_layout_folder, file_trans, set_amount, dead_vol_ul, save_location, save_file_name, specific_transfers=specific_transfers)

    save_file_name = "200_setes_230223"
    survey_folder = "C:/Users/phch/Desktop/more_data_files/full_plates"
    plate_layout_folder = "C:/Users/phch/Desktop/more_data_files/simulated_plate_layout"
    file_trans = "C:/Users/phch/Desktop/more_data_files/all_trans.xlsx"
    set_amount = 200
    dead_vol_ul = {"LDV": 2.5, "PP": 15}
    save_location = "C:/Users/phch/Desktop/more_data_files"



    new_worklist(survey_folder, plate_layout_folder, file_trans, set_amount, dead_vol_ul, save_location, save_file_name)


    # trans_report_controller(trans_data_folder, plate_layout_folder, all_trans_file, data_location, file_name, save_location)
    # well_report(all_trans_file)

    # print(file_names(path))

