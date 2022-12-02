from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl import Workbook, load_workbook
import xml.etree.ElementTree as ET
from math import floor

from get_data import well_compound_list, get_all_trans_data, get_survey_csv_data
from helper_func import folder_to_files


def _write_data_xml(compound_data_end, plates, save_filename):
    print("typing")

    wb = Workbook()
    ws = wb.active

    row = 1
    col = 1

    # write headlines:
    ws.cell(row=row, column=col, value="Plate")
    ws.cell(row=row, column=col + 1, value="source")
    ws.cell(row=row, column=col + 2, value="Compound")
    ws.cell(row=row, column=col + 3, value="Volume_needed(nL)")
    ws.cell(row=row, column=col + 4, value="Volume_left(uL)")
    ws.cell(row=row, column=col + 5, value="Amount_sets")
    ws.cell(row=row, column=col + 6, value="Well_amounts")

    row += 1
    temp_Set = ""

    for counter, Set in enumerate(plates):
        if temp_Set != Set:
            temp_Set = 1

        for barcode in plates[Set]:
            for compound in plates[Set][barcode]:
                for volume in plates[Set][barcode][compound]:
                    if plates[Set][barcode][compound][volume] != 0:
                        ws.cell(row=row, column=col, value=Set)
                        ws.cell(row=row, column=col + 1, value=barcode)
                        ws.cell(row=row, column=col + 2, value=compound)
                        vol_needed = plates[Set][barcode][compound][volume]
                        ws.cell(row=row, column=col + 3, value=vol_needed)
                        try:
                            compound_data_end[compound][barcode]
                        except KeyError:
                            ws.cell(row=row, column=col + 4, value="Not Found")
                        else:
                            for counter, wells in enumerate(compound_data_end[compound][barcode]):
                                vol_left = compound_data_end[compound][barcode][wells]
                                if wells == "total":
                                    ws.cell(row=row, column=col + 4, value=vol_left)
                                    set_amount = vol_left / (vol_needed / 1000)
                                    ws.cell(row=row, column=col + 5, value=set_amount)
                                else:
                                    ws.cell(row=row, column=col + 5 + counter, value=f"{wells}/{vol_left}")



                        row += 1

    wb.save(save_filename)
    print("done")


def _write_data_csv(compound_data_end, plates, save_filename):
    print(compound_data_end)
    wb = Workbook()
    ws = wb.active

    row = 1
    col = 1

    # write headlines:
    ws.cell(row=row, column=col, value="Plate")
    ws.cell(row=row, column=col + 1, value="source")
    ws.cell(row=row, column=col + 2, value="Compound")
    ws.cell(row=row, column=col + 3, value="Volume_needed(nL)")
    ws.cell(row=row, column=col + 4, value="Volume_left_total(uL)")
    ws.cell(row=row, column=col + 5, value="Amount_sets")
    ws.cell(row=row, column=col + 6, value="working_volume_left(uL)")
    ws.cell(row=row, column=col + 7, value="Amount_sets (based on working vol)")

    row += 1
    temp_Set = ""
    temp_total = {}
    temp_total_dead = {}
    temp_plate_spacer = {}
    ldv_dead_vol = 2.5
    for plate_set in plates:

        if temp_Set != plate_set:
            temp_Set = 1

        for barcode in plates[plate_set]:
            try:
                temp_total[barcode]
                temp_total_dead[barcode]
                temp_plate_spacer[barcode]
            except KeyError:
                temp_total[barcode] = {}
                temp_total_dead[barcode] = {}
                temp_plate_spacer[barcode] = {}

            for compound in plates[plate_set][barcode]:

                try:
                    temp_total[barcode][compound]
                    temp_total_dead[barcode][compound]
                    temp_plate_spacer[barcode][compound]
                except KeyError:
                    temp_total[barcode][compound] = 0
                    temp_total_dead[barcode][compound] = 0
                    temp_plate_spacer[barcode][compound] = 0

                for volume in plates[plate_set][barcode][compound]:
                    if plates[plate_set][barcode][compound][volume] != 0:
                        ws.cell(row=row, column=col, value=plate_set)
                        ws.cell(row=row, column=col + 1, value=barcode)
                        ws.cell(row=row, column=col + 2, value=compound)
                        vol_needed = plates[plate_set][barcode][compound][volume]
                        ws.cell(row=row, column=col + 3, value=vol_needed)
                        try:
                            compound_data_end[compound][barcode]
                        except KeyError:
                            ws.cell(row=row, column=col + 4, value="Not Found")
                        else:

                            for plate_name in compound_data_end[compound][barcode]:

                                w_col = col + 8 + temp_plate_spacer[barcode][compound]
                                ws.cell(row=row, column=w_col, value=plate_name).font = Font(bold=True)
                                ws.cell(row=row, column=w_col).fill = PatternFill(start_color='B284BE', end_color='B284BE',
                                                                         fill_type='solid')
                                for counter, wells in enumerate(compound_data_end[compound][barcode][plate_name]):
                                    vol_left = float(compound_data_end[compound][barcode][plate_name][wells])
                                    if wells == "total":
                                        temp_total[barcode][compound] += float(vol_left)
                                    else:
                                        vol_left -= ldv_dead_vol
                                        if vol_left < 0:
                                            vol_left = 0
                                        vol_left = round(vol_left, 2)
                                        temp_total_dead[barcode][compound] += float(vol_left)
                                        ws.cell(row=row, column=col + 8 + counter + temp_plate_spacer[barcode][compound]
                                                , value=f"{wells}/{vol_left}")
                                        if 0 < vol_left < 1:
                                            ws.cell(row=row,
                                                    column=col + 8 + counter + temp_plate_spacer[barcode][compound]
                                                    ).fill = PatternFill(start_color='FFFF00', end_color='FFFF00',
                                                                         fill_type='solid')

                                        elif vol_left == 0 or vol_left < 0:
                                            ws.cell(row=row,
                                                    column=col + 8 + counter + temp_plate_spacer[barcode][compound]
                                                    ).fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000',
                                                                         fill_type='solid')

                                temp_plate_spacer[barcode][compound] += counter + 1

                            ws.cell(row=row, column=col + 4, value=temp_total[barcode][compound])
                            set_amount = temp_total[barcode][compound] / (vol_needed / 1000)
                            set_amount = floor(set_amount)
                            ws.cell(row=row, column=col + 5, value=set_amount)
                            ws.cell(row=row, column=col + 6, value=temp_total_dead[barcode][compound])
                            set_amount_dead = temp_total_dead[barcode][compound] / (vol_needed / 1000)
                            set_amount_dead = floor(set_amount_dead)
                            ws.cell(row=row, column=col + 7, value=set_amount_dead)

                        row += 1

    wb.save(save_filename)


def _get_survey_xml_data(path):

    survey_data = {}

    file_list, _, _ = folder_to_files(path)

    for counter, files in enumerate(file_list):

        if files.split("\\")[-1].startswith("Survey"):
            # path = self.file_names(self.main_folder)
            doc = ET.parse(files)
            root = doc.getroot()

            # for counting plates and transferees
            for info in root.iter("platesurvey"):

                barcode = info.get("barcode")
                well_count = info.get("totalWells")
                date_time = info.get("date")
                time = date_time.split(" ")[-1].split(".")[0]
                temp_time = time
                try:
                    survey_data[barcode]
                except KeyError:
                    survey_data[barcode] = {"well_count": well_count, "well_info": {}}

            for well_info in root.iter("w"):
                well = well_info.get("n")
                volume = well_info.get("vl")
                fluid = well_info.get("fld")

                try:
                    survey_data[barcode]["well_info"][well]
                except KeyError:
                    survey_data[barcode]["well_info"][well] = {"volume": volume, "fluid": fluid}
                if counter > 0:
                    if temp_time > last_time:
                        survey_data[barcode]["well_info"][well]["volume"] = volume

                last_time = temp_time
    return survey_data


def _survey_to_compound_csv(well_data, survey_data):
    compound_data_end = {}

    for barcode in survey_data:
        for plate in survey_data[barcode]:
            for well in survey_data[barcode][plate]:
                try:
                    compound = well_data[plate][well]
                except KeyError:
                    compound = "None"
                vol_left = survey_data[barcode][plate][well]

                try:
                    compound_data_end[compound]
                except KeyError:
                    compound_data_end[compound] = {}

                try:
                    compound_data_end[compound][barcode]
                except KeyError:
                    compound_data_end[compound][barcode] = {}

                try:
                    compound_data_end[compound][barcode][plate]
                except KeyError:
                    compound_data_end[compound][barcode][plate] = {"total": 0, well: 0}

                compound_data_end[compound][barcode][plate]["total"] += float(vol_left)
                compound_data_end[compound][barcode][plate][well] = vol_left

    return compound_data_end


def _survey_to_compound_xml(well_data, survey_data):
    compound_data_end = {}

    for barcode in survey_data:
        for well in survey_data[barcode]["well_info"]:
            compound = well_data[well]
            vol_left = survey_data[barcode]["well_info"][well]["volume"]

            try:
                compound_data_end[compound]
            except KeyError:
                compound_data_end[compound] = {barcode: {"total": 0, well: 0}}

            compound_data_end[compound][barcode]["total"] += float(vol_left)
            compound_data_end[compound][barcode][well] = vol_left

    return compound_data_end


def survey_controller(survey_folder_csv, plate_layout_folder, save_file_name, save_location):
    # compound_list = folder_to_files(plate_layout_folder)
    file_save = f"{save_location}/{save_file_name}.xlsx"

    compound_data = well_compound_list(plate_layout_folder)

    _, single_set = get_all_trans_data(file_trans)
    survey_data_csv = get_survey_csv_data(survey_folder_csv)

    compound_data_end = _survey_to_compound_csv(compound_data, survey_data_csv)
    _write_data_csv(compound_data_end, single_set, file_save)


if __name__ == "__main__":

    survey_folder_xml = "C:/Users/phch/Desktop/more_data_files/2022-11-22"
    file_trans = "C:/Users/phch/Desktop/more_data_files/all_trans.xlsx"
    plate_layout_folder = "C:/Users/phch/Desktop/more_data_files/plate_layout"
    file_save_location = "C:/Users/phch/Desktop/more_data_files"
    survey_folder_csv = "C:/Users/phch/Desktop/more_data_files/surveys"
    # print(file_names(path))
    save_file_name = "survey_report"
    survey_controller(survey_folder_csv, plate_layout_folder, save_file_name, file_save_location)



