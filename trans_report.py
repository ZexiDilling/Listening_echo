from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl import Workbook, load_workbook
import xml.etree.ElementTree as ET


from get_data import get_xml_trans_data_printing_wells, well_compound_list


def report(wb, skipped_wells, skip_well_counter, working_list):
    ws = wb.active
    ws.title = "Overview_Report"
    row = 1
    col = 1

    # Headers
    ws.cell(row=row, column=col + 0, value="Source Plates").font = Font(bold=True)
    ws.cell(row=row, column=col + 1, value="Source Well").font = Font(bold=True)
    ws.cell(row=row, column=col + 2, value="Volume").font = Font(bold=True)
    ws.cell(row=row, column=col + 3, value="Destination Well").font = Font(bold=True)
    ws.cell(row=row, column=col + 4, value="Destination Plates").font = Font(bold=True)
    ws.cell(row=row, column=col + 5, value="Compound").font = Font(bold=True)
    ws.cell(row=row, column=col + 6, value="Comments").font = Font(bold=True)






def excel_controller(trans_data_folder, plate_layout_folder, data_location, file_name, save_location):
    trans_data = get_xml_trans_data_printing_wells(trans_data_folder)
    compound_data = well_compound_list(plate_layout_folder)

    return "Done"


if __name__ == "__main__":
    path = "C:/Users/phch/Desktop/echo_data"
    # print(file_names(path))

