from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl import Workbook, load_workbook
import xml.etree.ElementTree as ET

from helper_func import folder_to_files


def convert(file, save_location):
    """
    convert files
    :param file: the path to the file that needs to be converted
    :type file: str
    :param save_location: The location where the converted file will be saved
    :type save_location: str
    :return:
    """
    wb = load_workbook(file)
    ws = wb.active

    compound_data = {}

    for row_index, row in enumerate(ws.iter_rows()):

        if row_index > 0:
            for cell_index, cell in enumerate(row):
                if cell_index == 1 and cell.value:

                    temp_compound = cell.value

                if cell_index == 3 and cell.value:
                    print(cell.value)
                    try:
                        compound_data[temp_compound]
                    except KeyError:
                        compound_data[temp_compound] = cell.value.split(",")
                    else:
                        for values in cell.value.split(","):
                            compound_data[temp_compound].append(values)
    print(compound_data)
    wb = Workbook()
    ws = wb.active

    row = 1
    col = 1

    ws.cell(row=row, column=col, value="source_plate")
    ws.cell(row=row, column=col + 1, value="Drug")

    row += 1

    for compound in compound_data:
        for wells in compound_data[compound]:
            ws.cell(row=row, column=col, value=wells)
            ws.cell(row=row, column=col + 1, value=compound)
            row += 1
    wb.save(save_location)



if __name__ == "__main__":
    file = r"C:\Users\phch\OneDrive - Danmarks Tekniske Universitet\Mapper\Python_data\platePrinting\plate_print_P5_layout.xlsx"
    save_file = r"C:\Users\phch\OneDrive - Danmarks Tekniske Universitet\Mapper\Python_data\platePrinting\P5_ldv.xlsx"
    convert(file, save_file)