from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import csv
import xml.etree.ElementTree as ET
import re
import os

from helper_func import folder_to_files


def well_compound_list(path):
    """
    Takes excel file with wells in clm 1 and compound name in clm 2
    :param path: a folder with files
    :type path: str
    :return: compound_data - Data for what compound is in each well, based on excel files data.
    :rtype: dict
    """
    if os.path.isdir(path):
        file_list = folder_to_files(path)
    else:
        file_list = [path]

    compound_data = {}
    for files in file_list:
        plate_name = files.removesuffix(".xlsx").split("\\")[-1]
        # compound_data_org = {}
        plate_name = plate_name.replace("-", "_")
        compound_data[plate_name] = {}
        wb = load_workbook(filename=files)
        ws = wb.active
        for row, data in enumerate(ws):
            if row != 0:

                for col, cells in enumerate(data):

                    if col == 0:
                        temp_well = cells.value
                    if col == 1:
                        temp_compound = cells.value
                        try:
                            compound_data[plate_name][temp_well]
                        except KeyError:
                            compound_data[plate_name][temp_well] = temp_compound

    return compound_data


def get_all_trans_data(file_trans):
    """
    Takes excel data. all_transferees.
    :param file_trans: a file with all the transferees
    :type file_trans: str
    :return:
    all_plate_trans - all transferees, volume, wells and data for all transferees
    single_set - all transferees and data for a single set
    single_set_working_list - The information needed to create a workinglist, for a single set of plates, can be used
        for scaling up
    :rtype all_plate_trans: dict
    :rtype single_set: dict
    :rtype single_set_working_list: dict
    """
    all_plate_trans = {}
    is_first_set = True
    first_set = None
    single_set = {}
    single_set_working_list = {}

    wb = load_workbook(filename=file_trans)
    ws = wb.active

    for row, data in enumerate(ws):

        if row != 0:

            for col, cells in enumerate(data):

                if col == 0:
                    temp_comment = cells.value

                if col == 1:
                    temp_plate = cells.value
                    if row == 1:
                        first_set = temp_plate[:-2]

                    if temp_plate[:-2] != first_set:
                        is_first_set = False

                    temp_destination_plate = f"plate-{temp_plate[-1:]}"

                    try:
                        all_plate_trans[temp_plate[:-2]]
                    except KeyError:
                        all_plate_trans[temp_plate[:-2]] = {}

                    try:
                        all_plate_trans[temp_plate]
                    except KeyError:
                        all_plate_trans[temp_plate] = {}

                    if is_first_set:
                        try:
                            single_set[temp_plate[:-2]]
                        except KeyError:
                            single_set[temp_plate[:-2]] = {}

                if col == 2:
                    temp_dest_well = cells.value
                    temp_dest_well_list = re.split("(\d+)", temp_dest_well)
                    temp_dest_well = f"{temp_dest_well_list[0]}{int(temp_dest_well_list[1])}"

                if col == 3:
                    temp_compound = cells.value

                if col == 4:
                    temp_vol = float(cells.value)


                # if col == 5:
                #     temp_source_well = cells.value

                if col == 6:
                    temp_source_plate = cells.value
                    temp_source_plate = temp_source_plate.replace("-", "_")

                if col == 7:
                    temp_source_plate_type = cells.value

                    try:
                        all_plate_trans[temp_plate[:-2]][temp_source_plate]
                    except KeyError:
                        all_plate_trans[temp_plate[:-2]][temp_source_plate] = {}

                    try:
                        all_plate_trans[temp_plate][temp_source_plate]
                    except KeyError:
                        all_plate_trans[temp_plate][temp_source_plate] = {}

                    if is_first_set:
                        try:
                            single_set[temp_plate[:-2]][temp_source_plate]
                        except KeyError:
                            single_set[temp_plate[:-2]][temp_source_plate] = {}

            try:
                all_plate_trans[temp_plate][temp_source_plate][temp_compound]
            except KeyError:
                all_plate_trans[temp_plate][temp_source_plate][temp_compound] = {}
            try:
                all_plate_trans[temp_plate][temp_source_plate][temp_compound]["vol_needed"] += temp_vol
            except KeyError:
                all_plate_trans[temp_plate][temp_source_plate][temp_compound]["vol_needed"] = temp_vol

            try:
                all_plate_trans[temp_plate[:-2]][temp_source_plate][temp_compound]
            except KeyError:
                all_plate_trans[temp_plate[:-2]][temp_source_plate][temp_compound] = {}
            try:
                all_plate_trans[temp_plate[:-2]][temp_source_plate][temp_compound]["vol_needed"] += temp_vol
            except KeyError:
                all_plate_trans[temp_plate[:-2]][temp_source_plate][temp_compound]["vol_needed"] = temp_vol


            if is_first_set:
                try:
                    single_set[temp_plate[:-2]][temp_source_plate][temp_compound]
                except KeyError:
                    single_set[temp_plate[:-2]][temp_source_plate][temp_compound] = {}
                try:
                    single_set[temp_plate[:-2]][temp_source_plate][temp_compound]["vol_needed"] += temp_vol

                except KeyError:
                    single_set[temp_plate[:-2]][temp_source_plate][temp_compound]["vol_needed"] = temp_vol


                single_set_working_list[row] = {"destination_plate": temp_destination_plate,
                                                "destination_well": temp_dest_well,
                                                "compound": temp_compound,
                                                "volume_nl": temp_vol,
                                                "source_plate": temp_source_plate,
                                                "source_well": "",
                                                "sample_comment": temp_comment,
                                                "plate_type": temp_source_plate_type
                                                }

    return all_plate_trans, single_set, single_set_working_list


def get_survey_csv_data(path):
    """
    get survey data from CSV files - this is done by the echo from the main control modul. and is a standalone operation
    :param path: the path to the file
    :type path: str
    :return: survey_data - All the information from the survey
    :rtype: dict
    """
    survey_data = {}
    if os.path.isdir(path):
        file_list = folder_to_files(path)
    else:
        file_list = [path]

    for file in file_list:

        plate_name = file.split("\\")[-1].split(".")[0]
        plate_name = plate_name.replace("-", "_")
        if file.endswith(".csv"):
            if plate_name.split("_")[0] != "P23":
                barcode = "_".join(plate_name.split("_")[1:])
            else:
                barcode = "_".join(plate_name.split("_"))

            try:
                survey_data[barcode]
            except KeyError:
                survey_data[barcode] = {}

            try:
                survey_data[barcode][plate_name]
            except KeyError:
                survey_data[barcode][plate_name] = {}

            with open(file, newline="\n") as csv_file:
                data = csv.reader(csv_file, delimiter=",")
                for row_index, row in enumerate(data):
                    if row_index > 3:
                        for index, data in enumerate(row):
                            if index == 0:
                                col_letter = data
                            else:
                                temp_well = f"{col_letter}{index}"
                                if data:
                                    survey_data[barcode][plate_name][temp_well] = data
    return survey_data


def get_xml_trans_data_skipping_wells(path):
    """
    Looking through a transferee XML-file from the ECHO. and looking to see if there are any skipped wells for the
    transferee. If there are skipped wells.
    Skipped wells, are wells that have not been transferred due to different reasons.
    :param path: A path to the file. or a list of file names
    :type path: str or list
    :returns:
    all_data - All the data from the missing transferees
    skipped_wells - What wells are skipped
    skip_well_counter - Count how many times a single well is skipped
    working_list
    trans_plate_counter
    all_trans_counter
    :rtype all_data: list
    :rtype skipped_wells: dict
    :rtype skip_well_counter: int
    :rtype working_list: dict
    :rtype trans_plate_counter: list
    :rtype all_trans_counter: dict
    """

    # random data
    skipped_wells = {}
    working_list = {}
    all_data = []
    skip_well_counter = 0

    # counting Transferees
    trans_plate_counter = []
    counter_plates = []

    # counting transferees all data
    all_trans_counter = {}

    # Data for completed plates:
    zero_error_trans_plate = []
    error_trans_plate = []

    # checks if path is a directory

    if type(path) == list:
        file_list = path
    elif os.path.isdir(path):
        file_list = folder_to_files(path)
    else:
        file_list = [path]

    for files in file_list:

        if files.split("\\")[-1].startswith("Transfer"):
            # print(files)
            # path = self.file_names(self.main_folder)
            doc = ET.parse(files)
            root = doc.getroot()

            # for counting plates and transferees
            for plates in root.iter("plate"):
                barcode = plates.get("barcode")
                source_destination = plates.get("type")

                if source_destination == "destination":
                    counter_plates.append(barcode)
                    temp_d_barcode = barcode
                if source_destination == "source":
                    temp_s_barcode = barcode

            try:
                all_trans_counter[temp_d_barcode].append(temp_s_barcode)
            except KeyError:
                all_trans_counter[temp_d_barcode] = [temp_s_barcode]

            # find amount of well that is skipped
            for wells in root.iter("skippedwells"):
                wells_skipped = wells.get("total")
                if int(wells_skipped) != 0:
                    all_data.append(wells_skipped)
                    skip_well_counter += int(wells_skipped)

                    # finds barcode for source and destination
                    for dates in root.iter("transfer"):
                        date = dates.get("date")
                        all_data.append(date)

                    # finds barcode for source and destination
                    for plates in root.iter("plate"):
                        barcode = plates.get("barcode")
                        source_destination = plates.get("type")
                        all_data.append(source_destination + ", " + barcode)

                        if source_destination == "source":
                            temp_barcode = barcode
                            try:
                                skipped_wells[barcode]
                            except KeyError:
                                skipped_wells[barcode] = {}
                        if source_destination == "destination":
                            temp_dest_barcode = barcode
                            error_trans_plate.append(temp_dest_barcode)
                            try:
                                working_list[barcode]
                            except KeyError:
                                working_list[barcode] = {}

                            try:
                                working_list[barcode][temp_barcode]
                            except KeyError:
                                working_list[barcode][temp_barcode] = []

                    # finds destination and source wells data
                    for z in range(int(wells_skipped)):
                        temp_trans = []
                        dn = wells[z].get("dn")
                        n = wells[z].get("n")
                        reason = wells[z].get("reason")
                        vt = wells[z].get("vt")
                        all_data.append("SW: " + n + " DW: " + dn + " vol: " + vt)
                        all_data.append(" reason: " + reason)
                        temp_trans.append(n)
                        temp_trans.append(float(vt))
                        temp_trans.append(dn)
                        # Gets only the error code from reason
                        reason = reason.split(":")[0]
                        try:
                            skipped_wells[temp_barcode][n]["counter"] += 1
                            skipped_wells[temp_barcode][n]["vol"] += float(vt)
                            skipped_wells[temp_barcode][n]["reason"].append(reason)
                        except KeyError:
                            skipped_wells[temp_barcode][n] = {"counter": 1, "vol": float(vt), "reason": [reason]}

                        working_list[temp_dest_barcode][temp_barcode].append(temp_trans)
                else:
                    # finds barcode for destination
                    for plates in root.iter("plate"):
                        barcode = plates.get("barcode")
                        source_destination = plates.get("type")

                        if source_destination == "destination":
                            zero_error_trans_plate.append(barcode)

    # counting plates
    # counts the number of repeated barcodes and makes a list with the barcode and amount of
    # instance with the same name

    for plates in counter_plates:
        number = counter_plates.count(plates)
        trans_plate_counter.append(str(plates) + "," + str(number))

    # remove duplicates from the list
    trans_plate_counter = list(dict.fromkeys(trans_plate_counter))
    # print(skip_well_counter)
    plates_to_remove = []
    if zero_error_trans_plate:
        for plates in zero_error_trans_plate:
            if plates in error_trans_plate:
                plates_to_remove.append(plates)

        for plates in plates_to_remove:
            zero_error_trans_plate.remove(plates)

    return all_data, skipped_wells, skip_well_counter, working_list, trans_plate_counter, all_trans_counter, \
           zero_error_trans_plate, temp_d_barcode


def get_xml_trans_data_printing_wells(path):
    """
    Getting all the data for successful transferees
    :param path: the path to the file
    :type path: str
    :return: all_trans_data - All data for the transferee
    :rtype: dict
    """
    all_trans_data = {}
    if type(path) == list:
        file_list = path
    else:
        file_list = folder_to_files(path)

    for files in file_list:
        if files.split("\\")[-1].startswith("Transfer"):
            trans_name = files.split("\\")[-1].removesuffix(".xml")
            # path = self.file_names(self.main_folder)
            doc = ET.parse(files)
            root = doc.getroot()
            # find amount of well that is skipped
            for wells in root.iter("printmap"):
                wells_printed = wells.get("total")
                if int(wells_printed) != 0:

                    all_trans_data[trans_name] = {"destination_plate": "", "source_plate": "", "transferees": [],
                                                         "date": ""}

                    for plates in root.iter("plate"):

                        barcode = plates.get("barcode")
                        source_destination = plates.get("type")

                        if source_destination == "source":
                            all_trans_data[trans_name]["source_plate"] = barcode

                        if source_destination == "destination":
                            all_trans_data[trans_name]["destination_plate"] = barcode

                    for dates in root.iter("transfer"):
                        all_trans_data[trans_name]["date"] = dates.get("date")

                    # finds destination and source wells data
                    for z in range(int(wells_printed)):

                        dest_well = wells[z].get("dn")
                        source_well = wells[z].get("n")
                        volume = wells[z].get("vt")
                        temp_trans = {"destination_well": dest_well, "source_well": source_well, "volume": volume}

                        all_trans_data[trans_name]["transferees"].append(temp_trans)
    return all_trans_data


def get_plate_lay_from_excel_raw(file):
    """
    Takes excel file, with compound name in col_x, vol in col_y, wells in col_z
    :param file: The excel file
    :type file: str
    :return: dict over compound in wells:
    :rtype: dict
    """

    wb = load_workbook(file)
    ws = wb.active
    compound = {}

    for line_index, line in enumerate(ws):
        if line_index != 0:
            for cell_index, cell in enumerate(line):
                if cell_index == 0:
                    drug = cell.value

                if drug:
                    try:
                        compound[drug]
                    except KeyError:
                        compound[drug] = []
                    if cell_index == 3:
                        well_values = cell.value
                        temp_well_list = well_values.split(",")
                        for wells in temp_well_list:
                            try:
                                start_well, end_well = wells.split("-")
                            except ValueError:
                                compound[drug].append(wells)
                            else:
                                letter, start_num, _ = re.split('(\d+)', start_well.strip())
                                end_num = re.split('(\d+)', end_well.strip())[1]
                                for i in range(int(start_num), int(end_num) + 1):
                                    well = f"{letter}{i}"
                                    compound[drug].append(well)

    return compound


def convert_compound_data_to_plate_layout(compound_data, full_path):
    """
    A dict of compound, and wells
    :param compound_data:
    :return:
    """
    wb = Workbook()
    ws = wb.active
    row = 1
    col = 1

    # Headling
    ws.cell(row=row, column=col, value="source_plate").font = Font(bold=True)
    ws.cell(row=row, column=col + 1, value="compound").font = Font(bold=True)

    row += 1

    for compound in compound_data:
        for well in compound_data[compound]:
            ws.cell(row=row, column=col, value=well)
            ws.cell(row=row, column=col + 1, value=compound)
            row += 1

    wb.save(full_path)

if __name__ == "__main__":
    location = "C:/Users/phch/Desktop/more_data_files"
    file = "2022-12-09-Compound_200sets_with well position.xlsx"
    full_path = f"{location}/{file}"
    file_layout = "ldv_200.xlsx"
    full_path_layout = f"{location}/{file_layout}"
    compound = get_plate_lay_from_excel_raw(full_path)
    convert_compound_data_to_plate_layout(compound, full_path_layout)