from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl import Workbook

from get_data import get_xml_trans_data_skipping_wells


def write_to_excel_plate_transferees_list_of_plates(wb, data):

    ws = wb.create_sheet('Plate_trans_list')

    # write headers
    ws.cell(row=1, column=1, value="destination plate").font = Font(bold=True)
    ws.merge_cells("B1:F1")
    ws["B1"].value = "source plates"
    ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B1"].font = Font(bold=True)

    col = 1
    row = 2

    for destination in data:
        ws.cell(row=row, column=col, value=destination)
        for index, source in enumerate(data[destination]):
            ws.cell(row=row, column=col + 1 + index, value=source)

        row += 1


def write_to_excel_plate_transferees(wb, data):

    ws = wb.create_sheet('Plate_trans_counter')

    # split the data up by ","
    temp = []
    for i in enumerate(data):
        temp.append(i[1].split(","))

    # write headers
    ws.cell(row=1, column=1, value="destination plate").font = Font(bold=True)
    ws.cell(row=1, column=2, value="counter").font = Font(bold=True)

    # write in the data, barcodes for destination plates and the number of transferees.
    col = 1
    row = 2
    for plate, counts in temp:
        ws.cell(row=row, column=col, value=plate)
        ws.cell(row=row, column=col + 1, value=counts)
        row += 1


def write_to_excel_error_report(wb, data, type):
    # Error_Report
    # all data
    ws = wb.create_sheet(f"{type}")
    row = 1
    col = 1
    i = 0
    while i < len(data):
        # write to excel worksheet
        ws.cell(row=row, column=col, value=data[i+1]).font = Font(bold=True)    # date
        ws.cell(row=row, column=col + 1, value=data[i + 2]).font = Font(bold=True)  # source barcode
        ws.cell(row=row, column=col + 2, value=data[i + 3]).font = Font(bold=True)  # destination barcode
        # sets n = to wells skipped
        n = data[i]
        ws.cell(row=row, column=col + 3, value=n).font = Font(bold=True)    # writes amount of wells skipped
        # sets counter for next loop
        temp = i + 4
        i = temp
        q = 0

        # loop for writing out wells
        for k in range(temp, (int(n)*2)+temp, 2):
            # makes sure that the Excel file is only 5-row wide
            if q == 5:
                row = row + 2
                q = 0
            # writes SW, DW and vol for col 1, writes reason for skipped in col 2
            ws.cell(row=row + 1, column=col + q, value=data[k])
            ws.cell(row=row + 2, column=col + q, value=data[k + 1])
            q += 1

        # counter sets for next iteration
        row = row + 3
        col = col
        i = int(n)*2+i


def work_list(wb, data, type):
    ws = wb.create_sheet(f'{type}_worklist')

    row = 1
    col = 1

    # write headlines
    ws.cell(row=row, column=col, value="source_plate")
    ws.cell(row=row, column=col + 1, value="source_well")
    ws.cell(row=row, column=col + 2, value="trans_vol")
    ws.cell(row=row, column=col + 3, value="destination_well")
    ws.cell(row=row, column=col + 4, value="destination_plate")
    row += 1

    for destination in data:
        for source in data[destination]:
            for index, trans in enumerate(data[destination][source]):
                # source plate
                ws.cell(row=row, column=col, value=source)

                for off_set, info in enumerate(data[destination][source][index]):
                    # source well trans vol, destination will
                    ws.cell(row=row, column=col + off_set + 1, value=str(info))
                # destination plate
                ws.cell(row=row, column=col + 4, value=destination)
                row += 1


def report(wb, skipped_wells, skip_well_counter, working_list):
    ws = wb.active
    ws.title = "Overview_Report"
    row = 1
    col = 1

    source_plate_amount = 0
    destination_plate_amount = 0
    source_plate_list = []
    destination_plate_list = []
    for destination_plates in working_list:
        destination_plate_amount += 1
        destination_plate_list.append(destination_plates)
        for source_plate in working_list[destination_plates]:

            if source_plate not in source_plate_list:
                source_plate_list.append(source_plate)
                source_plate_amount += 1

    # overview
    ws.cell(row=row, column=col, value="Skipped Wells counter").font = Font(bold=True)
    ws.cell(row=row, column=col + 1, value=skip_well_counter)
    row += 1
    ws.cell(row=row, column=col, value="Source Plates counter").font = Font(bold=True)
    ws.cell(row=row, column=col + 1, value=source_plate_amount)
    row += 1
    ws.cell(row=row, column=col, value="Destination plate counter").font = Font(bold=True)
    ws.cell(row=row, column=col + 1, value=destination_plate_amount)
    row += 2
    ws.cell(row=row, column=col, value="Destination Plates").font = Font(bold=True)
    ws.cell(row=row, column=col + 1, value="Source Plates").font = Font(bold=True)
    row += 1

    temp_row = row
    for destination in destination_plate_list:
        ws.cell(row=temp_row, column=col, value=destination)
        temp_row += 1

    temp_row = row
    for source in source_plate_list:
        ws.cell(row=temp_row, column=col + 1, value=source)
        temp_row += 1

    temp_row = row
    temp_col = col
    for source_plate in skipped_wells:
        ws.cell(row=temp_row - 1, column=temp_col + 3, value="Source Plates").font = Font(bold=True)
        ws.cell(row=temp_row - 1, column=temp_col + 4, value="Source Wells").font = Font(bold=True)
        ws.cell(row=temp_row - 1, column=temp_col + 5, value="Well Instance").font = Font(bold=True)
        ws.cell(row=temp_row - 1, column=temp_col + 6, value="Well Volume").font = Font(bold=True)

        ws.cell(row=temp_row, column=temp_col + 3, value=source_plate)
        for index, source_well in enumerate(skipped_wells[source_plate]):
            ws.cell(row=temp_row + index, column=temp_col + 4, value=source_well)
            ws.cell(row=temp_row + index, column=temp_col + 5, value=skipped_wells[source_plate][source_well]["counter"])
            ws.cell(row=temp_row + index, column=temp_col + 6, value=skipped_wells[source_plate][source_well]["vol"])

        temp_col += 4


def excel_controller(data_location, file_name, save_location):
    wb = Workbook()
    all_data, skipped_wells, skip_well_counter, working_list, trans_plate_counter, all_trans_counter = \
        get_xml_trans_data_skipping_wells(data_location)
    report(wb, skipped_wells, skip_well_counter, working_list)
    write_to_excel_plate_transferees_list_of_plates(wb, all_trans_counter)
    write_to_excel_plate_transferees(wb, trans_plate_counter)
    write_to_excel_error_report(wb, all_data, "Error_Report")
    work_list(wb, working_list, "Old")

    wb.save(f"{save_location}/{file_name}.xlsx")
    return "Done"

if __name__ == "__main__":
    path = "C:/Users/phch/Desktop/echo_data"
    # print(file_names(path))
    excel_controller(path)



